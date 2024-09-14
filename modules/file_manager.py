import os
import xlsxwriter
import openpyxl
import shutil
import logging
from PyQt5.QtCore import QMutex

class FileManager:
    """文件管理器"""

    def __init__(self):
        self.auto_save_file = "all_questions_log.xlsx"
        self.backup_file = "all_questions_log_backup.xlsx"
        self.save_buffer = []
        self.mutex = QMutex()
        if not os.path.exists(self.auto_save_file):
            self.create_excel_file()

    def create_excel_file(self):
        """创建 Excel 文件并写入标题行"""
        workbook = xlsxwriter.Workbook(self.auto_save_file)
        worksheet = workbook.add_worksheet()

        headers = ["题目", "答案", "反馈", "生成时间", "回答时间", "反馈时间"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        workbook.close()

    def add_to_save_buffer(self, question, answer, feedback, q_time, a_time, f_time):
        """将数据添加到保存队列中"""
        self.mutex.lock()
        try:
            save_data = {
                "question": question,
                "answer": answer,
                "feedback": feedback,
                "question_time": q_time,
                "answer_time": a_time,
                "feedback_time": f_time,
            }
            self.save_buffer.append(save_data)
        finally:
            self.mutex.unlock()

    def save_to_file(self):
        """保存数据到 Excel 文件"""
        workbook = openpyxl.load_workbook(self.auto_save_file)
        sheet = workbook.active

        row_start = sheet.max_row + 1
        for data in self.save_buffer:
            row = row_start
            sheet.cell(row=row, column=1, value=data["question"])
            sheet.cell(row=row, column=2, value=data["answer"])
            sheet.cell(row=row, column=3, value=data["feedback"])
            sheet.cell(row=row, column=4, value=data["question_time"])
            sheet.cell(row=row, column=5, value=data["answer_time"])
            sheet.cell(row=row, column=6, value=data["feedback_time"])
            row_start += 1

        workbook.save(self.auto_save_file)

    def backup_data(self):
        """备份数据文件"""
        try:
            shutil.copy(self.auto_save_file, self.backup_file)
            logging.info("数据备份成功")
        except Exception as e:
            logging.error(f"数据备份失败: {e}")

    def read_files(self, file_names):
        """读取上传的文件内容"""
        file_content = ""
        for file_name in file_names:
            if not os.path.exists(file_name):
                continue
            try:
                if file_name.endswith(".txt"):
                    with open(file_name, "r", encoding="utf-8") as f:
                        file_content += f.read() + "\n"
                elif file_name.endswith(".docx"):
                    import docx

                    doc = docx.Document(file_name)
                    file_content += "\n".join([para.text for para in doc.paragraphs]) + "\n"
                elif file_name.endswith(".pdf"):
                    from PyPDF2 import PdfReader

                    reader = PdfReader(file_name)
                    for page in reader.pages:
                        file_content += page.extract_text() + "\n"
                else:
                    # 尝试以文本方式读取未知类型的文件
                    with open(file_name, "r", encoding="utf-8", errors="ignore") as f:
                        file_content += f.read() + "\n"
            except Exception as e:
                logging.error(f"无法读取文件 {file_name}: {e}")
                continue
        return file_content

    def export_to_file(self, file_path):
        """导出数据到指定的 Excel 文件"""
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()

        headers = ["题目", "答案", "反馈", "生成时间", "回答时间", "反馈时间"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        workbook2 = openpyxl.load_workbook(self.auto_save_file)
        sheet2 = workbook2.active

        for row_idx, row in enumerate(sheet2.iter_rows(min_row=2), start=1):
            for col_idx, cell in enumerate(row):
                worksheet.write(row_idx, col_idx, cell.value)

        workbook.close()
